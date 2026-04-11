"""
Gerador do Demonstrativo do Fluxo de Caixa (DFC).

Modulo irmao de reports/dre.py. Consome pagamentos reais de account.payment
e produz duas abas adicionais no mesmo Workbook de Demonstrativos Financeiros:

- "DFC {ano}": fluxo de caixa mensal com SALDO ACUMULADO e linhas separadas
  por cartao de credito (diferidas conforme closing_day/due_day).
- "Detalhamento Fluxo de Caixa": row-by-row de cada pagamento considerado,
  com data original, data efetiva (apos diferimento) e observacoes.

Diferente da aba "DRE Caixa" (que e competencia reagrupada por
invoice_date_due), este relatorio olha o que efetivamente entrou/saiu
do caixa/banco — e, para cartao de credito, projeta o fechamento da
fatura para quando ela realmente e paga.

Config esperada no clients/<slug>.yaml:

    cash_flow:
      saldo_inicial_por_ano:
        2026: 0.00
      credit_cards:
        - name: "Cartao Itau Black"
          id: null          # opcional, prevalece sobre name quando presente
          closing_day: 25
          due_day: 5

Sem cash_flow ou sem credit_cards, a aba DFC simplesmente nao e criada
(o arquivo Excel sai apenas com as abas de DRE).
"""

import os
from calendar import monthrange
from datetime import date, datetime

from openpyxl.utils import get_column_letter

# Reutiliza estilos e helpers do DRE — mantem consistencia visual entre
# as abas DRE e DFC do mesmo arquivo.
from reports.dre import (
    MESES,
    FMT_BRL,
    _S_TITULO,
    _S_SUBTITULO,
    _S_HEADER,
    _S_SECAO,
    _S_DADO,
    _S_TOTAL,
    _S_RESULT,
    _S_FINAL,
    _S_OBS,
    _F_EFETIVO,
    _F_PROVISORIO,
    _fill_row,
    _write_data_row,
    _write_formula_row,
    obs_from_states,
)

from openpyxl.styles import Alignment


# ─── Loader de config ────────────────────────────────────────────────────────


def load_cash_flow_config(slug: str | None = None) -> dict:
    """Le clients/<slug>.yaml e retorna o dict cash_flow (ou vazio).

    slug = None usa a variavel de ambiente ODOO_CLIENT. Se nem o env var
    nem o arquivo existirem, devolve {} silenciosamente — quem chama
    interpreta como "nao gerar DFC".
    """
    slug = slug or os.environ.get("ODOO_CLIENT")
    if not slug:
        return {}

    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(repo_root, "clients", f"{slug}.yaml")
    if not os.path.exists(path):
        return {}

    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        raise ImportError("pyyaml nao instalado. Rode: pip install pyyaml")

    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return data.get("cash_flow") or {}


# ─── Helpers de data ────────────────────────────────────────────────────────


def _parse_date(value) -> date | None:
    """Converte valor do Odoo (string YYYY-MM-DD ou date) em datetime.date."""
    if not value:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def _defer_card_payment_date(
    charge_date: date,
    closing_day: int,
    due_day: int,
) -> date:
    """Data em que a fatura do cartao contendo esta compra sera paga.

    Regra:
    - Se a compra ocorrer ate (inclusive) o closing_day, ela entra na
      fatura que fecha no closing_day deste mes e vence no due_day do
      proximo mes.
    - Se ocorrer depois do closing_day, entra na fatura do proximo mes,
      paga no due_day do mes subsequente.

    due_day alem do ultimo dia do mes (ex: 31 em fevereiro) e clampado
    para o ultimo dia do mes usando calendar.monthrange.
    """
    if charge_date.day <= closing_day:
        settle_month = charge_date.month + 1
        settle_year = charge_date.year
    else:
        settle_month = charge_date.month + 2
        settle_year = charge_date.year
    while settle_month > 12:
        settle_month -= 12
        settle_year += 1
    last_day = monthrange(settle_year, settle_month)[1]
    return date(settle_year, settle_month, min(due_day, last_day))


# ─── Busca e agregacao ──────────────────────────────────────────────────────


_PAYMENT_FIELDS = [
    "id",
    "name",
    "date",
    "amount",
    "amount_company_currency_signed",
    "payment_type",
    "journal_id",
    "partner_id",
    "state",
    "ref",
]


def _resolve_card_journals(odoo, cards_cfg: list, log=None) -> dict:
    """Resolve a lista de cartoes configurada em ids de account.journal.

    Retorna {journal_id: {"name", "closing_day", "due_day"}}. Levanta
    ValueError com lista de candidatos se um nome de cartao resolver para
    mais de um diario.
    """
    def _emit(msg):
        if log:
            log(msg)

    result: dict = {}
    for card in cards_cfg or []:
        if "closing_day" not in card or "due_day" not in card:
            raise ValueError(
                f"cash_flow.credit_cards: cartao '{card.get('name') or card.get('id')}' "
                "precisa de closing_day e due_day."
            )
        jid = card.get("id")
        name = card.get("name")
        if jid:
            hit = odoo.search(
                "account.journal",
                [["id", "=", int(jid)]],
                fields=["id", "name"],
                limit=1,
            )
            if not hit:
                raise ValueError(f"cash_flow: diario id={jid} nao encontrado")
            resolved_name = hit[0]["name"]
        elif name:
            hits = odoo.search(
                "account.journal",
                [["name", "ilike", name]],
                fields=["id", "name"],
                limit=None,
            )
            if len(hits) == 0:
                raise ValueError(
                    f"cash_flow: nenhum account.journal bateu com '{name}'"
                )
            if len(hits) > 1:
                candidates = ", ".join(f"{h['id']}={h['name']}" for h in hits)
                raise ValueError(
                    f"cash_flow: nome '{name}' ambiguo ({len(hits)} matches): {candidates}. "
                    "Informe 'id' no YAML para desambiguar."
                )
            jid = hits[0]["id"]
            resolved_name = hits[0]["name"]
        else:
            raise ValueError(
                "cash_flow.credit_cards: cada cartao precisa de 'id' ou 'name'."
            )
        result[int(jid)] = {
            "name": resolved_name,
            "closing_day": int(card["closing_day"]),
            "due_day": int(card["due_day"]),
        }
        _emit(f"  Cartao resolvido: [{jid}] {resolved_name} "
              f"(close {card['closing_day']} / due {card['due_day']})")
    return result


def _signed_amount(payment: dict) -> float:
    """Valor do pagamento em moeda da empresa, signed.

    Prefere amount_company_currency_signed (Odoo 16+). Fallback: amount
    com sinal derivado de payment_type (inbound=+, outbound=-).
    """
    val = payment.get("amount_company_currency_signed")
    if val is not None and val != 0:
        return float(val)
    raw = float(payment.get("amount") or 0.0)
    if payment.get("payment_type") == "outbound":
        return -raw
    return raw


def buscar_dados_fluxo_caixa(
    odoo,
    ano: int,
    cash_flow_config: dict,
    log=None,
) -> dict:
    """Busca pagamentos do Odoo e agrega o fluxo de caixa do ano.

    Args:
        odoo: OdooClient autenticado.
        ano: Ano fiscal alvo (ex: 2026).
        cash_flow_config: dict da chave cash_flow do clients/<slug>.yaml.
            Esperado: {"saldo_inicial_por_ano": {ano: float},
                       "credit_cards": [...]}
        log: callable opcional para progresso (ex: print).

    Returns:
        dict com entradas/saidas mensais, agregacoes por cartao,
        ajustes de ano anterior/posterior, saldo inicial e lista
        detalhada de pagamentos para a aba de auditoria.
    """
    def _emit(msg: str):
        if log:
            log(msg)

    _emit(f"Buscando dados de fluxo de caixa ({ano})...")

    cards_cfg = (cash_flow_config or {}).get("credit_cards") or []
    card_map = _resolve_card_journals(odoo, cards_cfg, log=log) if cards_cfg else {}
    card_journal_ids = set(card_map.keys())

    saldo_inicial = float(
        ((cash_flow_config or {}).get("saldo_inicial_por_ano") or {}).get(ano)
        or 0.0
    )

    data_ini = f"{ano}-01-01"
    data_fim = f"{ano}-12-31"

    _emit(f"  Buscando account.payment entre {data_ini} e {data_fim}...")
    pagamentos = odoo.search(
        "account.payment",
        [
            ["state", "in", ["posted", "paid"]],
            ["date", ">=", data_ini],
            ["date", "<=", data_fim],
        ],
        fields=_PAYMENT_FIELDS,
        limit=None,
    )
    _emit(f"    {len(pagamentos)} pagamento(s) do ano")

    # Lookback do ano anterior — so nos diarios de cartao, para trazer
    # compras de nov/dez do ano anterior cuja fatura vence em jan/fev do ano atual.
    pagamentos_lookback: list = []
    if card_journal_ids:
        lookback_ini = f"{ano - 1}-11-01"
        lookback_fim = f"{ano - 1}-12-31"
        _emit(f"  Lookback cartao: {lookback_ini} .. {lookback_fim}")
        pagamentos_lookback = odoo.search(
            "account.payment",
            [
                ["journal_id", "in", list(card_journal_ids)],
                ["state", "in", ["posted", "paid"]],
                ["date", ">=", lookback_ini],
                ["date", "<=", lookback_fim],
            ],
            fields=_PAYMENT_FIELDS,
            limit=None,
        )
        _emit(f"    {len(pagamentos_lookback)} pagamento(s) no lookback")

    entradas_operacionais = {m: 0.0 for m in range(1, 13)}
    saidas_operacionais = {m: 0.0 for m in range(1, 13)}
    entradas_states: set = set()
    saidas_states: set = set()

    saidas_cartao_por_card: dict = {
        jid: {
            "nome": info["name"],
            "closing_day": info["closing_day"],
            "due_day": info["due_day"],
            "por_mes": {m: 0.0 for m in range(1, 13)},
            "proximo_ano": 0.0,
            "states": set(),
        }
        for jid, info in card_map.items()
    }
    ajuste_cartao_ano_anterior = {m: 0.0 for m in range(1, 13)}
    ajuste_states: set = set()

    pagamentos_detalhe: list = []

    def _process(payment: dict, is_lookback: bool):
        data_raw = payment.get("date")
        dt = _parse_date(data_raw)
        if not dt:
            return
        journal = payment.get("journal_id") or [None, ""]
        jid = journal[0] if isinstance(journal, list) else journal
        journal_name = journal[1] if isinstance(journal, list) and len(journal) > 1 else ""
        partner = payment.get("partner_id") or [None, ""]
        partner_name = partner[1] if isinstance(partner, list) and len(partner) > 1 else ""
        state = payment.get("state") or ""
        signed = _signed_amount(payment)

        # Cartao: sempre diferido.
        if jid in card_map:
            card = card_map[jid]
            data_efetiva = _defer_card_payment_date(dt, card["closing_day"], card["due_day"])
            valor = abs(signed)  # cartao e sempre saida
            obs_detalhe = ""
            if data_efetiva.year == ano:
                if is_lookback:
                    # Veio de nov/dez do ano anterior, fatura cai em ano atual
                    ajuste_cartao_ano_anterior[data_efetiva.month] += valor
                    ajuste_states.add(state)
                    obs_detalhe = f"Diferido de {dt.isoformat()} (ano anterior)"
                else:
                    bucket = saidas_cartao_por_card[jid]
                    bucket["por_mes"][data_efetiva.month] += valor
                    bucket["states"].add(state)
                    obs_detalhe = f"Diferido: compra em {dt.isoformat()}"
            elif data_efetiva.year > ano and not is_lookback:
                # Compra do ano atual cuja fatura vence no ano seguinte
                bucket = saidas_cartao_por_card[jid]
                bucket["proximo_ano"] += valor
                bucket["states"].add(state)
                obs_detalhe = f"Fatura vence em {data_efetiva.isoformat()} (proximo ano)"
            else:
                # Lookback com data efetiva fora do ano -> ignorar
                return

            pagamentos_detalhe.append({
                "data_original": dt.isoformat(),
                "data_efetiva": data_efetiva.isoformat(),
                "tipo": "Cartao",
                "cartao": card["name"],
                "journal": journal_name,
                "partner": partner_name,
                "ref": payment.get("ref") or "",
                "pagamento": payment.get("name") or "",
                "valor": -valor,  # cartao sempre sai do caixa
                "obs": obs_detalhe,
                "state": state,
            })
            return

        # Nao-cartao: so conta pagamentos do ano (nao os do lookback).
        if is_lookback:
            return
        mes = dt.month
        tipo = payment.get("payment_type") or ""
        if tipo == "inbound" or signed > 0:
            entradas_operacionais[mes] += abs(signed)
            entradas_states.add(state)
            pagamentos_detalhe.append({
                "data_original": dt.isoformat(),
                "data_efetiva": dt.isoformat(),
                "tipo": "Entrada",
                "cartao": "",
                "journal": journal_name,
                "partner": partner_name,
                "ref": payment.get("ref") or "",
                "pagamento": payment.get("name") or "",
                "valor": abs(signed),
                "obs": "",
                "state": state,
            })
        else:
            saidas_operacionais[mes] += abs(signed)
            saidas_states.add(state)
            pagamentos_detalhe.append({
                "data_original": dt.isoformat(),
                "data_efetiva": dt.isoformat(),
                "tipo": "Saida",
                "cartao": "",
                "journal": journal_name,
                "partner": partner_name,
                "ref": payment.get("ref") or "",
                "pagamento": payment.get("name") or "",
                "valor": -abs(signed),
                "obs": "",
                "state": state,
            })

    for p in pagamentos:
        _process(p, is_lookback=False)
    for p in pagamentos_lookback:
        _process(p, is_lookback=True)

    # Ordena detalhe por data efetiva
    pagamentos_detalhe.sort(key=lambda r: (r["data_efetiva"], r["data_original"]))

    return {
        "ano": ano,
        "hoje": str(date.today()),
        "saldo_inicial": saldo_inicial,
        "entradas_operacionais": entradas_operacionais,
        "saidas_operacionais": saidas_operacionais,
        "entradas_obs": obs_from_states(entradas_states),
        "saidas_obs": obs_from_states(saidas_states),
        "saidas_cartao_por_card": saidas_cartao_por_card,
        "ajuste_cartao_ano_anterior": ajuste_cartao_ano_anterior,
        "ajuste_cartao_obs": obs_from_states(ajuste_states),
        "pagamentos_detalhe": pagamentos_detalhe,
        "total_pagamentos_processados": len(pagamentos) + len(pagamentos_lookback),
    }


# ─── Aba: DFC ───────────────────────────────────────────────────────────────


def _build_fluxo_caixa_sheet(ws, ano: int, dados_fc: dict, moeda: str, empresa: str):
    """Constroi a aba DFC {ano} no workbook.

    Layout imita a aba DRE: titulo, subtitulo, header, secoes coloridas,
    formulas mensais (Jan-Dez) + TOTAL (col O) com SUM, e um SALDO
    ACUMULADO com running sum celula a celula.
    """
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 12
    for ci in range(3, 16):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    # Row 1: Titulo
    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = f"{empresa} \u2014 DFC {ano}"
    c.font = _S_TITULO["font"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    _fill_row(ws, 1, _S_TITULO["fill"])

    # Row 2: Subtitulo
    ws.merge_cells("A2:O2")
    c = ws["A2"]
    c.value = (
        f"Extraido em {dados_fc['hoje']} | Valores em {moeda} | "
        "Base: account.payment (posted/paid) | "
        "Cartoes diferidos conforme closing_day/due_day"
    )
    c.font = _S_SUBTITULO["font"]
    c.alignment = Alignment(horizontal="center")
    _fill_row(ws, 2, _S_SUBTITULO["fill"])

    # Row 4: Headers
    headers = ["Descri\u00e7\u00e3o", "Obs"] + MESES + [f"TOTAL {ano}"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _S_HEADER["font"]
        c.fill = _S_HEADER["fill"]
        c.alignment = Alignment(horizontal="center")

    # ── ENTRADAS ──
    ws.cell(row=6, column=1, value="(+) ENTRADAS OPERACIONAIS").font = _S_SECAO["font"]
    _fill_row(ws, 6, _S_SECAO["fill"])

    row_recebimentos = 7
    ws.cell(row=row_recebimentos, column=1,
            value="Recebimentos (pagamentos inbound)")
    _write_data_row(ws, row_recebimentos,
                    dados_fc["entradas_operacionais"],
                    dados_fc["entradas_obs"], _S_DADO)

    row_tot_entradas = 8
    ws.cell(row=row_tot_entradas, column=1, value="TOTAL ENTRADAS")
    _write_formula_row(ws, row_tot_entradas,
                       f"{{c}}{row_recebimentos}", _S_TOTAL)

    # ── SAIDAS ──
    ws.cell(row=10, column=1, value="(-) SAIDAS OPERACIONAIS").font = _S_SECAO["font"]
    _fill_row(ws, 10, _S_SECAO["fill"])

    row_cursor = 11
    ws.cell(row=row_cursor, column=1,
            value="Pagamentos a fornecedores (nao-cartao)")
    _write_data_row(ws, row_cursor,
                    dados_fc["saidas_operacionais"],
                    dados_fc["saidas_obs"], _S_DADO)
    saida_rows = [row_cursor]
    row_cursor += 1

    # Uma linha por cartao
    cartao_items = list(dados_fc["saidas_cartao_por_card"].items())
    for jid, info in cartao_items:
        ws.cell(row=row_cursor, column=1,
                value=f"{info['nome']} (diferido)")
        _write_data_row(ws, row_cursor, info["por_mes"],
                        obs_from_states(info["states"]), _S_DADO)
        saida_rows.append(row_cursor)
        row_cursor += 1

    # Linha de ajuste do ano anterior
    ws.cell(row=row_cursor, column=1,
            value="Ajuste: faturas de cartao do ano anterior")
    _write_data_row(ws, row_cursor,
                    dados_fc["ajuste_cartao_ano_anterior"],
                    dados_fc["ajuste_cartao_obs"], _S_DADO)
    saida_rows.append(row_cursor)
    row_cursor += 1

    # TOTAL SAIDAS (soma com sinal negativo)
    row_cursor += 1  # linha em branco
    row_tot_saidas = row_cursor
    tmpl = "-(" + "+".join(f"{{c}}{r}" for r in saida_rows) + ")"
    ws.cell(row=row_tot_saidas, column=1, value="TOTAL SAIDAS")
    _write_formula_row(ws, row_tot_saidas, tmpl, _S_TOTAL)
    row_cursor += 1

    # ── FLUXO LIQUIDO ──
    row_cursor += 1  # linha em branco
    row_fluxo = row_cursor
    ws.cell(row=row_fluxo, column=1, value="FLUXO LIQUIDO DO MES")
    _write_formula_row(ws, row_fluxo,
                       f"{{c}}{row_tot_entradas}+{{c}}{row_tot_saidas}",
                       _S_RESULT)
    row_cursor += 2

    # ── SALDO ──
    row_saldo_ini = row_cursor
    ws.cell(row=row_saldo_ini, column=1, value="SALDO INICIAL")
    c = ws.cell(row=row_saldo_ini, column=3,
                value=round(dados_fc["saldo_inicial"], 2))
    c.font = _S_DADO["font"]
    c.number_format = FMT_BRL
    row_cursor += 1

    row_saldo = row_cursor
    ws.cell(row=row_saldo, column=1, value="SALDO ACUMULADO")
    # Estilo em todas as colunas C..O
    _fill_row(ws, row_saldo, _S_FINAL["fill"])
    for ci in range(1, 16):
        ws.cell(row=row_saldo, column=ci).font = _S_FINAL["font"]

    # Running sum: C = saldo_ini + fluxo(jan)
    c = ws.cell(row=row_saldo, column=3,
                value=f"=C{row_saldo_ini}+C{row_fluxo}")
    c.number_format = FMT_BRL
    c.font = _S_FINAL["font"]
    c.fill = _S_FINAL["fill"]
    # D..N: previous + current month's fluxo
    for ci in range(4, 15):
        cl = get_column_letter(ci)
        pl = get_column_letter(ci - 1)
        c = ws.cell(row=row_saldo, column=ci,
                    value=f"={pl}{row_saldo}+{cl}{row_fluxo}")
        c.number_format = FMT_BRL
        c.font = _S_FINAL["font"]
        c.fill = _S_FINAL["fill"]
    # O (TOTAL) = saldo final = valor da coluna N
    c = ws.cell(row=row_saldo, column=15, value=f"=N{row_saldo}")
    c.number_format = FMT_BRL
    c.font = _S_FINAL["font"]
    c.fill = _S_FINAL["fill"]
    row_cursor += 2

    # ── Rodape: faturas diferidas para o proximo ano ──
    total_prox_ano = sum(
        info["proximo_ano"] for info in dados_fc["saidas_cartao_por_card"].values()
    )
    if total_prox_ano > 0:
        row_footer = row_cursor
        ws.cell(
            row=row_footer,
            column=1,
            value=(
                f"Faturas de cartao diferidas para {ano + 1} "
                f"(nao entram em {ano})"
            ),
        )
        c = ws.cell(row=row_footer, column=15, value=round(total_prox_ano, 2))
        c.number_format = FMT_BRL
        c.font = _S_OBS

    ws.freeze_panes = "C5"


# ─── Aba: Detalhamento Fluxo de Caixa ───────────────────────────────────────


def _build_detalhe_fluxo_sheet(wb, ano: int, dados_fc: dict, moeda: str):
    """Aba auxiliar row-by-row com todos os pagamentos processados."""
    ws = wb.create_sheet(title="Detalhamento Fluxo de Caixa")

    n_cols = 10
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"Detalhamento do Fluxo de Caixa \u2014 {ano}"
    c.font = _S_TITULO["font"]
    c.fill = _S_TITULO["fill"]
    c.alignment = Alignment(horizontal="center")
    _fill_row(ws, 1, _S_TITULO["fill"], 1, n_cols)

    headers = [
        "Data Original", "Data Efetiva", "Tipo", "Cartao",
        "Diario", "Contraparte", "Refer\u00eancia", "Pagamento",
        f"Valor ({moeda})", "Observa\u00e7\u00e3o",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _S_HEADER["font"]
        c.fill = _S_HEADER["fill"]
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([13, 13, 10, 22, 22, 28, 22, 16, 16, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for linha in dados_fc["pagamentos_detalhe"]:
        ws.cell(row=row, column=1, value=linha["data_original"])
        ws.cell(row=row, column=2, value=linha["data_efetiva"])
        tipo = linha["tipo"]
        tc = ws.cell(row=row, column=3, value=tipo)
        if tipo == "Entrada":
            tc.font = _F_EFETIVO
        elif tipo in ("Saida", "Cartao"):
            tc.font = _F_PROVISORIO
        ws.cell(row=row, column=4, value=linha["cartao"])
        ws.cell(row=row, column=5, value=linha["journal"])
        ws.cell(row=row, column=6, value=linha["partner"])
        ws.cell(row=row, column=7, value=linha["ref"])
        ws.cell(row=row, column=8, value=linha["pagamento"])
        c = ws.cell(row=row, column=9, value=round(linha["valor"], 2))
        c.number_format = FMT_BRL
        ws.cell(row=row, column=10, value=linha["obs"])
        row += 1

    ws.freeze_panes = "A3"
