"""
Gerador de DRE (Demonstracao do Resultado do Exercicio) em Excel (.xlsx).

Produz planilha profissional com 3 abas:
- "DRE {ano}": demonstrativo mensal com formulas Excel reais
- "Detalhamento Receitas": todas as faturas de venda por mes
- "Detalhamento Despesas": todas as faturas de compra categorizadas

Agnostico de empresa — nome da empresa e mapeamento de fornecedores para
categorias de despesa sao recebidos como parametros.

Moeda: todos os valores na moeda da empresa (amount_total_signed).
Faturas em moedas estrangeiras sao convertidas pelo Odoo pela taxa do dia.

Uso CLI:
    python reports/dre.py 2026
    python reports/dre.py 2026 --output /tmp/dre_2026.xlsx
    python reports/dre.py 2026 --mapeamento data/mapeamento.json
"""

import argparse
import json
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Constantes ──────────────────────────────────────────────────────────────

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
         "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

CATEGORIAS_DESPESA = [
    "Pessoal / Serviços Profissionais",
    "Terceirização / Subcontratação",
    "Projetos Especiais",
    "Software / SaaS / Infraestrutura",
    "Impostos e Taxas",
    "Outras Despesas",
]

CATEGORIA_DEFAULT = CATEGORIAS_DESPESA[-1]


# ─── Helpers de dados ────────────────────────────────────────────────────────

def _mes_de(data: str | None, fallback: str | None = None) -> int | None:
    """Extrai o mês (1-12) de uma data YYYY-MM-DD. Usa fallback se data for None."""
    d = data or fallback
    return int(d.split("-")[1]) if d else None


def _split_analytic_key(key: str) -> list[int]:
    """Decompõe chaves compostas de analytic_distribution ('35,99,22') em IDs individuais."""
    return [int(x) for x in key.split(",") if x.strip().isdigit()]

# ─── Estilos ─────────────────────────────────────────────────────────────────

FMT_BRL = '#,##0.00;(#,##0.00);"-"'
FMT_PCT = "0.0%"

_S_TITULO = {
    "font": Font(name="Arial", size=14, bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="1F3864"),
}
_S_SUBTITULO = {
    "font": Font(name="Arial", size=9, italic=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="2F5496"),
}
_S_HEADER = {
    "font": Font(name="Arial", size=12, bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="2F5496"),
}
_S_SECAO = {
    "font": Font(name="Arial", size=10, bold=True),
    "fill": PatternFill("solid", fgColor="E2EFDA"),
}
_S_DADO = {
    "font": Font(name="Arial", size=10),
}
_S_TOTAL = {
    "font": Font(name="Arial", size=10, bold=True),
    "fill": PatternFill("solid", fgColor="FFF2CC"),
}
_S_RESULT = {
    "font": Font(name="Arial", size=11, bold=True, color="1F3864"),
    "fill": PatternFill("solid", fgColor="BDD7EE"),
}
_S_FINAL = {
    "font": Font(name="Arial", size=12, bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="1F3864"),
}
_S_MARGEM = {
    "font": Font(name="Arial", size=9, italic=True, color="666666"),
}
_S_OBS = Font(name="Arial", size=9, color="666666")

_F_EFETIVO = Font(name="Arial", size=10, color="006100")
_F_PROVISORIO = Font(name="Arial", size=10, color="C65911")


# ─── Funcoes publicas ────────────────────────────────────────────────────────


def categorizar_por_conta(
    analitica_nome: str,
    conta_nome: str,
    mapeamento: dict[str, list[str]],
) -> str:
    """Categoriza despesa pela conta analitica (prioritario) ou conta contabil (fallback).

    Compara os nomes de forma case-insensitive e por substring.

    Args:
        analitica_nome: Nome da conta analitica resolvida (vazio se nao houver distribuicao).
        conta_nome: Nome da conta do plano de contas (account_id[1] do Odoo).
        mapeamento: {categoria: [termos]} para classificar. Se vazio, retorna CATEGORIA_DEFAULT.

    Returns:
        Nome da categoria de despesa.
    """
    def _match(nome: str) -> str | None:
        n = nome.lower()
        for cat, termos in mapeamento.items():
            if any(t.lower() in n for t in termos):
                return cat
        return None

    if analitica_nome:
        cat = _match(analitica_nome)
        if cat:
            return cat
    if conta_nome:
        cat = _match(conta_nome)
        if cat:
            return cat
    return CATEGORIA_DEFAULT


def obs_from_states(states: set) -> str:
    """Retorna 'Efetivo', 'Provisorio' ou 'Efet+Prov' com base nos estados das faturas."""
    if not states:
        return ""
    has_posted = "posted" in states
    has_draft = "draft" in states
    if has_posted and has_draft:
        return "Efet+Prov"
    if has_posted:
        return "Efetivo"
    if has_draft:
        return "Provisorio"
    return ""


def gerar_excel_dre(*, ano, hoje, empresa, moeda,
                    receita, imposto, despesa,
                    receita_obs, imposto_obs, despesa_obs,
                    receita_caixa, imposto_caixa, despesa_caixa,
                    receita_caixa_obs, imposto_caixa_obs, despesa_caixa_obs,
                    vendas, linhas_compra, output_path):
    """Gera o arquivo Excel completo do DRE e salva em output_path.

    Produz 4 abas:
    - DRE {ano} — Competencia: agregado por invoice_date
    - DRE {ano} — Caixa: agregado por invoice_date_due
    - Detalhamento Receitas: faturas de venda com ambas as datas
    - Detalhamento Despesas: linhas de compra com ambas as datas

    Returns:
        Caminho do arquivo salvo (str).
    """
    wb = Workbook()
    ws_comp = wb.active
    ws_comp.title = f"DRE {ano} \u2014 Compet\u00eancia"
    _build_dre_sheet(ws_comp, ano, hoje, empresa, moeda,
                     receita, imposto, despesa,
                     receita_obs, imposto_obs, despesa_obs,
                     sufixo="Compet\u00eancia", regime_desc="data da fatura (invoice_date)")
    ws_caixa = wb.create_sheet(title=f"DRE {ano} \u2014 Caixa")
    _build_dre_sheet(ws_caixa, ano, hoje, empresa, moeda,
                     receita_caixa, imposto_caixa, despesa_caixa,
                     receita_caixa_obs, imposto_caixa_obs, despesa_caixa_obs,
                     sufixo="Caixa", regime_desc="data de vencimento (invoice_date_due)")
    _build_receitas_sheet(wb, ano, vendas, moeda)
    _build_despesas_sheet(wb, ano, linhas_compra, moeda)
    wb.save(output_path)
    return output_path


# ─── Helpers internos ────────────────────────────────────────────────────────


def _fill_row(ws, row, fill, col_start=1, col_end=15):
    for ci in range(col_start, col_end + 1):
        ws.cell(row=row, column=ci).fill = fill


def _write_data_row(ws, row, data, obs, style):
    """Escreve dados mensais (C-N) + SUM no TOTAL (O)."""
    font = style["font"]
    fill = style.get("fill")
    ws.cell(row=row, column=1).font = font
    if obs:
        ws.cell(row=row, column=2, value=obs).font = _S_OBS
    if fill:
        _fill_row(ws, row, fill)
    for m in range(1, 13):
        c = ws.cell(row=row, column=m + 2, value=round(data[m], 2))
        c.font = font
        c.number_format = FMT_BRL
    c = ws.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")
    c.font = font
    c.number_format = FMT_BRL


def _write_formula_row(ws, row, tmpl, style, fmt=FMT_BRL):
    """Escreve formula para colunas C-O. tmpl usa {c} como placeholder."""
    font = style["font"]
    fill = style.get("fill")
    ws.cell(row=row, column=1).font = font
    if fill:
        _fill_row(ws, row, fill)
    for ci in range(3, 16):
        cl = get_column_letter(ci)
        c = ws.cell(row=row, column=ci, value="=" + tmpl.replace("{c}", cl))
        c.font = font
        c.number_format = fmt
        if fill:
            c.fill = fill


# ─── Aba 1: DRE ─────────────────────────────────────────────────────────────


def _build_dre_sheet(ws, ano, hoje, empresa, moeda, receita, imposto, despesa,
                     receita_obs, imposto_obs, despesa_obs,
                     sufixo="Compet\u00eancia", regime_desc="data da fatura"):

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 12
    for ci in range(3, 16):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    # Row 1: Titulo
    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = f"{empresa} \u2014 DRE {ano} \u2014 Regime de {sufixo}"
    c.font = _S_TITULO["font"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    _fill_row(ws, 1, _S_TITULO["fill"])

    # Row 2: Subtitulo
    ws.merge_cells("A2:O2")
    c = ws["A2"]
    c.value = (
        f"Extra\u00eddo do Odoo em {hoje} | Valores em {moeda} | "
        f"Compet\u00eancia por {regime_desc} | "
        "Moedas estrangeiras convertidas pela taxa do Odoo na data da fatura"
    )
    c.font = _S_SUBTITULO["font"]
    c.alignment = Alignment(horizontal="center")
    _fill_row(ws, 2, _S_SUBTITULO["fill"])

    # Row 4: Headers
    headers = ["Descri\u00e7\u00e3o", "Obs"] + MESES + [f"TOTAL {ano}"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _S_HEADER["font"]
        c.fill = _S_HEADER["fill"]
        c.alignment = Alignment(horizontal="center")

    # ── RECEITA ──
    ws.cell(row=6, column=1, value="RECEITA OPERACIONAL BRUTA").font = _S_SECAO["font"]
    _fill_row(ws, 6, _S_SECAO["fill"])

    ws.cell(row=7, column=1, value="Receita de Servi\u00e7os (Faturamento)")
    _write_data_row(ws, 7, receita, receita_obs, _S_DADO)

    ws.cell(row=8, column=1, value="TOTAL RECEITA BRUTA")
    _write_formula_row(ws, 8, "{c}7", _S_TOTAL)

    # ── DEDUCOES ──
    ws.cell(row=10, column=1, value="(-) DEDU\u00c7\u00d5ES DA RECEITA").font = _S_SECAO["font"]
    _fill_row(ws, 10, _S_SECAO["fill"])

    ws.cell(row=11, column=1, value="(-) Impostos sobre Servi\u00e7os")
    _write_data_row(ws, 11, imposto, imposto_obs, _S_DADO)

    ws.cell(row=12, column=1, value="TOTAL DEDU\u00c7\u00d5ES")
    _write_formula_row(ws, 12, "-{c}11", _S_TOTAL)

    # ── RECEITA LIQUIDA ──
    ws.cell(row=14, column=1, value="RECEITA OPERACIONAL L\u00cdQUIDA")
    _write_formula_row(ws, 14, "{c}8+{c}12", _S_RESULT)

    # ── CSP ──
    ws.cell(row=16, column=1, value="(-) CUSTOS DOS SERVI\u00c7OS PRESTADOS").font = _S_SECAO["font"]
    _fill_row(ws, 16, _S_SECAO["fill"])

    ws.cell(row=17, column=1, value="(-) Pessoal / Servi\u00e7os Profissionais")
    _write_data_row(ws, 17,
                    despesa["Pessoal / Servi\u00e7os Profissionais"],
                    despesa_obs["Pessoal / Servi\u00e7os Profissionais"],
                    _S_DADO)

    ws.cell(row=18, column=1, value="(-) Terceiriza\u00e7\u00e3o / Subcontrata\u00e7\u00e3o")
    _write_data_row(ws, 18,
                    despesa["Terceiriza\u00e7\u00e3o / Subcontrata\u00e7\u00e3o"],
                    despesa_obs["Terceiriza\u00e7\u00e3o / Subcontrata\u00e7\u00e3o"],
                    _S_DADO)

    ws.cell(row=19, column=1, value="TOTAL CSP")
    _write_formula_row(ws, 19, "-({c}17+{c}18)", _S_TOTAL)

    # ── LUCRO BRUTO ──
    ws.cell(row=21, column=1, value="LUCRO BRUTO")
    _write_formula_row(ws, 21, "{c}14+{c}19", _S_RESULT)

    ws.cell(row=22, column=1, value="Margem Bruta %").font = _S_MARGEM["font"]
    _write_formula_row(ws, 22, "IF({c}14=0,0,{c}21/{c}14)", _S_MARGEM, FMT_PCT)

    # ── DESPESAS OPERACIONAIS ──
    ws.cell(row=24, column=1, value="(-) DESPESAS OPERACIONAIS").font = _S_SECAO["font"]
    _fill_row(ws, 24, _S_SECAO["fill"])

    ws.cell(row=25, column=1, value="(-) Software / SaaS / Infraestrutura")
    _write_data_row(ws, 25,
                    despesa["Software / SaaS / Infraestrutura"],
                    despesa_obs["Software / SaaS / Infraestrutura"],
                    _S_DADO)

    ws.cell(row=26, column=1, value="(-) Impostos e Taxas")
    _write_data_row(ws, 26,
                    despesa["Impostos e Taxas"],
                    despesa_obs["Impostos e Taxas"],
                    _S_DADO)

    ws.cell(row=28, column=1, value="TOTAL DESPESAS OPERACIONAIS")
    _write_formula_row(ws, 28, "-({c}25+{c}26)", _S_TOTAL)

    # ── EBITDA ──
    ws.cell(row=30, column=1, value="EBITDA (Resultado Operacional)")
    _write_formula_row(ws, 30, "{c}21+{c}28", _S_RESULT)

    ws.cell(row=31, column=1, value="Margem EBITDA %").font = _S_MARGEM["font"]
    _write_formula_row(ws, 31, "IF({c}14=0,0,{c}30/{c}14)", _S_MARGEM, FMT_PCT)

    # ── RESULTADO ──
    ws.cell(row=33, column=1, value="RESULTADO L\u00cdQUIDO DO EXERC\u00cdCIO")
    _write_formula_row(ws, 33, "{c}30", _S_FINAL)

    ws.cell(row=34, column=1, value="Margem L\u00edquida %").font = _S_MARGEM["font"]
    _write_formula_row(ws, 34, "IF({c}14=0,0,{c}33/{c}14)", _S_MARGEM, FMT_PCT)

    ws.freeze_panes = "C5"


# ─── Aba 2: Detalhamento Receitas ────────────────────────────────────────────


def _build_receitas_sheet(wb, ano, vendas, moeda):
    ws = wb.create_sheet(title="Detalhamento Receitas")

    n_cols = 10
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"Detalhamento de Receitas \u2014 {ano}"
    c.font = _S_TITULO["font"]
    c.fill = _S_TITULO["fill"]
    c.alignment = Alignment(horizontal="center")
    _fill_row(ws, 1, _S_TITULO["fill"], 1, n_cols)

    headers = [
        "Compet\u00eancia", "Vencimento", "Cliente", "Moeda Original",
        f"L\u00edquido ({moeda})", f"Impostos ({moeda})", f"Total ({moeda})",
        "Status", "Fatura", "M\u00eas Compet\u00eancia",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _S_HEADER["font"]
        c.fill = _S_HEADER["fill"]
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([13, 13, 32, 14, 16, 16, 16, 12, 18, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for v in vendas:
        data_comp = v.get("invoice_date") or ""
        data_venc = v.get("invoice_date_due") or data_comp
        mes_num = int(data_comp.split("-")[1]) if data_comp else 1
        moeda_orig = v["currency_id"][1] if v["currency_id"] else moeda
        total_brl = v["amount_total_signed"]
        untaxed_brl = v.get("amount_untaxed_signed", total_brl)
        tax_brl = total_brl - untaxed_brl
        status = "Efetivo" if v["state"] == "posted" else "Provis\u00f3rio"

        ws.cell(row=row, column=1, value=data_comp)
        ws.cell(row=row, column=2, value=data_venc)
        ws.cell(row=row, column=3, value=v["partner_id"][1] if v["partner_id"] else "")
        ws.cell(row=row, column=4, value=moeda_orig)

        for col, val in [(5, untaxed_brl), (6, tax_brl), (7, total_brl)]:
            c = ws.cell(row=row, column=col, value=round(val, 2))
            c.number_format = FMT_BRL

        sc = ws.cell(row=row, column=8, value=status)
        sc.font = _F_EFETIVO if v["state"] == "posted" else _F_PROVISORIO
        ws.cell(row=row, column=9, value=v["name"])
        ws.cell(row=row, column=10, value=MESES[mes_num - 1])
        row += 1

    ws.freeze_panes = "A3"


# ─── Aba 3: Detalhamento Despesas ─────────────────────────────────────────


def _build_despesas_sheet(wb, ano, linhas_compra, moeda):
    ws = wb.create_sheet(title="Detalhamento Despesas")

    n_cols = 10
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = f"Detalhamento de Despesas \u2014 {ano}"
    c.font = _S_TITULO["font"]
    c.fill = _S_TITULO["fill"]
    c.alignment = Alignment(horizontal="center")
    _fill_row(ws, 1, _S_TITULO["fill"], 1, n_cols)

    headers = [
        "Compet\u00eancia", "Vencimento", "Fornecedor", "Fatura", "Categoria",
        "Conta Anal\u00edtica", "Conta Cont\u00e1bil", f"Valor ({moeda})", "Status",
        "M\u00eas Compet\u00eancia",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _S_HEADER["font"]
        c.fill = _S_HEADER["fill"]
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([13, 13, 26, 18, 28, 22, 28, 16, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for linha in linhas_compra:
        status_posted = linha["status"] == "posted"
        ws.cell(row=row, column=1, value=linha.get("data_competencia", ""))
        ws.cell(row=row, column=2, value=linha.get("data_vencimento", ""))
        ws.cell(row=row, column=3, value=linha["fornecedor"])
        ws.cell(row=row, column=4, value=linha["fatura"])
        ws.cell(row=row, column=5, value=linha["categoria"])
        ws.cell(row=row, column=6, value=linha["analitica"])
        ws.cell(row=row, column=7, value=linha["conta"])
        c = ws.cell(row=row, column=8, value=round(linha["valor"], 2))
        c.number_format = FMT_BRL
        sc = ws.cell(row=row, column=9, value="Efetivo" if status_posted else "Provis\u00f3rio")
        sc.font = _F_EFETIVO if status_posted else _F_PROVISORIO
        ws.cell(row=row, column=10, value=MESES[linha["mes"] - 1])
        row += 1

    ws.freeze_panes = "A3"


# ─── Busca e agregação (usada pelo CLI e pelo MCP) ───────────────────────────


def buscar_dados_dre(odoo, ano: int, mapeamento: dict, log=None) -> dict:
    """Busca faturas no Odoo e agrega os dados para o DRE.

    Args:
        odoo: Instância autenticada de OdooClient.
        ano: Ano fiscal.
        mapeamento: {categoria: [termos]} para classificar despesas.
        log: Callable opcional para saída de progresso (ex: print).

    Returns:
        Dict com todos os dados necessários para gerar_excel_dre() + metadados
        (empresa, moeda, contagens).
    """
    def _emit(msg: str):
        if log:
            log(msg)

    hoje = str(date.today())

    _emit("Buscando dados da empresa...")
    company = odoo.search("res.company", [], fields=["name", "currency_id"], limit=1)
    empresa = company[0]["name"] if company else "Empresa"
    moeda_raw = company[0]["currency_id"][1] if company and company[0].get("currency_id") else "BRL"
    moeda = moeda_raw.split()[0].strip("[]") if moeda_raw else "BRL"
    _emit(f"  Empresa: {empresa} | Moeda: {moeda}")

    data_ini = f"{ano}-01-01"
    data_fim = f"{ano}-12-31"
    campos_fatura = [
        "name", "invoice_date", "invoice_date_due", "move_type", "state",
        "partner_id", "currency_id",
        "amount_total_signed", "amount_untaxed_signed",
    ]

    _emit(f"Buscando faturas de venda ({ano})...")
    vendas = odoo.search("account.move", [
        ["move_type", "=", "out_invoice"],
        ["state", "in", ["posted", "draft"]],
        ["invoice_date", ">=", data_ini],
        ["invoice_date", "<=", data_fim],
    ], fields=campos_fatura, limit=None)
    _emit(f"  {len(vendas)} fatura(s) encontrada(s)")

    _emit(f"Buscando faturas de compra ({ano})...")
    compras = odoo.search("account.move", [
        ["move_type", "=", "in_invoice"],
        ["state", "in", ["posted", "draft"]],
        ["invoice_date", ">=", data_ini],
        ["invoice_date", "<=", data_fim],
    ], fields=campos_fatura, limit=None)
    _emit(f"  {len(compras)} fatura(s) encontrada(s)")

    receita = {m: 0.0 for m in range(1, 13)}
    imposto = {m: 0.0 for m in range(1, 13)}
    receita_caixa = {m: 0.0 for m in range(1, 13)}
    imposto_caixa = {m: 0.0 for m in range(1, 13)}
    receita_states: set = set()
    receita_caixa_states: set = set()
    for v in vendas:
        if not v.get("invoice_date"):
            continue
        total = v.get("amount_total_signed") or 0.0
        untaxed = v.get("amount_untaxed_signed") or total
        mes = _mes_de(v.get("invoice_date"))
        if mes:
            receita[mes] += untaxed
            imposto[mes] += total - untaxed
            receita_states.add(v["state"])
        mes_c = _mes_de(v.get("invoice_date_due"), v.get("invoice_date"))
        if mes_c:
            receita_caixa[mes_c] += untaxed
            imposto_caixa[mes_c] += total - untaxed
            receita_caixa_states.add(v["state"])

    compras_by_id = {c["id"]: c for c in compras}
    compra_ids = list(compras_by_id.keys())

    linhas_odoo = []
    if compra_ids:
        _emit(f"Buscando linhas de despesa ({len(compra_ids)} faturas)...")
        linhas_odoo = odoo.search("account.move.line", [
            ["move_id", "in", compra_ids],
            ["display_type", "not in", ["line_section", "line_note"]],
            ["tax_line_id", "=", False],
            ["price_subtotal", "!=", 0],
        ], fields=[
            "move_id", "name", "account_id", "analytic_distribution", "price_subtotal",
        ], limit=None)
        _emit(f"  {len(linhas_odoo)} linha(s) encontrada(s)")

    analytic_ids: set = set()
    for l in linhas_odoo:
        if l.get("analytic_distribution"):
            for k in l["analytic_distribution"].keys():
                analytic_ids.update(_split_analytic_key(k))

    analiticas: dict = {}
    if analytic_ids:
        _emit(f"Resolvendo {len(analytic_ids)} conta(s) analitica(s)...")
        aa = odoo.search("account.analytic.account", [
            ["id", "in", list(analytic_ids)],
        ], fields=["id", "name"], limit=None)
        analiticas = {r["id"]: r["name"] for r in aa}

    despesa = {cat: {m: 0.0 for m in range(1, 13)} for cat in CATEGORIAS_DESPESA}
    despesa_caixa = {cat: {m: 0.0 for m in range(1, 13)} for cat in CATEGORIAS_DESPESA}
    despesa_states: dict = {cat: set() for cat in CATEGORIAS_DESPESA}
    despesa_caixa_states: dict = {cat: set() for cat in CATEGORIAS_DESPESA}
    linhas_compra = []

    for l in linhas_odoo:
        move_id = l["move_id"][0] if isinstance(l["move_id"], list) else l["move_id"]
        compra = compras_by_id.get(move_id)
        if not compra or not compra.get("invoice_date"):
            continue
        data_comp = compra.get("invoice_date") or ""
        data_venc = compra.get("invoice_date_due") or data_comp
        mes = _mes_de(data_comp)
        mes_c = _mes_de(data_venc, data_comp)
        if not mes:
            continue
        analitica_nome = ""
        if l.get("analytic_distribution"):
            top_key = max(l["analytic_distribution"], key=lambda k: l["analytic_distribution"][k])
            ids = _split_analytic_key(top_key)
            analitica_nome = analiticas.get(ids[0], "") if ids else ""
        conta_nome = l["account_id"][1] if l.get("account_id") else ""
        cat = categorizar_por_conta(analitica_nome, conta_nome, mapeamento)
        if cat not in despesa:
            cat = CATEGORIA_DEFAULT
        valor = abs(l.get("price_subtotal") or 0.0)
        despesa[cat][mes] += valor
        despesa_states[cat].add(compra["state"])
        if mes_c:
            despesa_caixa[cat][mes_c] += valor
            despesa_caixa_states[cat].add(compra["state"])
        linhas_compra.append({
            "mes": mes,
            "data_competencia": data_comp,
            "data_vencimento": data_venc,
            "fornecedor": compra["partner_id"][1] if compra.get("partner_id") else "",
            "fatura": compra.get("name", ""),
            "categoria": cat,
            "analitica": analitica_nome,
            "conta": conta_nome,
            "valor": valor,
            "status": compra["state"],
        })

    return {
        "hoje": hoje,
        "empresa": empresa,
        "moeda": moeda,
        "vendas": vendas,
        "linhas_compra": linhas_compra,
        "receita": receita,
        "imposto": imposto,
        "receita_caixa": receita_caixa,
        "imposto_caixa": imposto_caixa,
        "despesa": despesa,
        "despesa_caixa": despesa_caixa,
        "receita_obs": obs_from_states(receita_states),
        "imposto_obs": obs_from_states(receita_states),
        "despesa_obs": {cat: obs_from_states(despesa_states[cat]) for cat in CATEGORIAS_DESPESA},
        "receita_caixa_obs": obs_from_states(receita_caixa_states),
        "imposto_caixa_obs": obs_from_states(receita_caixa_states),
        "despesa_caixa_obs": {cat: obs_from_states(despesa_caixa_states[cat]) for cat in CATEGORIAS_DESPESA},
    }


# ─── CLI ─────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Gera DRE em Excel a partir do Odoo")
    parser.add_argument("ano", type=int, help="Ano fiscal (ex: 2026)")
    parser.add_argument("--output", "-o", help="Caminho do .xlsx (default: reports/dre_{ano}.xlsx)")
    parser.add_argument("--mapeamento", "-m", help="JSON com mapeamento de categorias")
    args = parser.parse_args()

    mapeamento = {}
    if args.mapeamento:
        with open(args.mapeamento, encoding="utf-8") as f:
            mapeamento = json.load(f)

    output_path = args.output or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "reports", f"dre_{args.ano}.xlsx",
    )

    print(f"\n=== DRE {args.ano} ===")
    print(f"Output: {output_path}\n")

    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from odoo import OdooClient  # noqa: PLC0415

    odoo = OdooClient()
    print(f"Conectado ao Odoo (uid={odoo.uid})", flush=True)

    dados = buscar_dados_dre(odoo, args.ano, mapeamento, log=lambda m: print(m, flush=True))

    print("Gerando planilha Excel...", flush=True)
    path = gerar_excel_dre(
        ano=args.ano, output_path=output_path, **{k: dados[k] for k in dados if k != "hoje"},
        hoje=dados["hoje"],
    )

    receita_total = sum(dados["receita"].values())
    despesa_total = sum(sum(v.values()) for v in dados["despesa"].values())
    moeda = dados["moeda"]
    print(f"\n=== Concluido ===")
    print(f"Arquivo:           {path}")
    print(f"Empresa:           {dados['empresa']} | Moeda: {moeda}")
    print(f"Faturas de venda:  {len(dados['vendas'])}")
    print(f"Linhas de despesa: {len(dados['linhas_compra'])}")
    print(f"Receita bruta:     {moeda} {receita_total:,.2f}")
    print(f"Despesas:          {moeda} {despesa_total:,.2f}")
    print(f"Resultado liquido: {moeda} {receita_total - despesa_total:,.2f}")


if __name__ == "__main__":
    main()
